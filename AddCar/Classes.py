from datetime import date
from http import client
from multiprocessing.sharedctypes import Value
from PyQt5 import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import csv
from openpyxl import load_workbook
from win32com import client
import os
import subprocess

class Voiture:
    def __init__(self,matricule,marque,couleur,etat,date,prix):
        self.matricule=matricule
        self.marque=marque
        self.couleur=couleur
        self.etat=etat
        self.date=date
        self.prix=prix


class Client:
    def __init__(self,cin,nom,prenom,age,adresse,mail,tel):
        self.cin=cin
        self.nom=nom
        self.prenom=prenom
        self.age=age
        self.adresse=adresse
        self.mail=mail
        self.tel=tel

class Location:
    def __init__(self,numero,cin,matricule,date,duree,montant):
        self.numero=numero
        self.cin=cin
        self.matricule=matricule
        self.date=date
        self.duree=duree
        self.montant=montant

class Main:
    def __init__(self):
        self.dictClient=dict()
        self.dictVoiture=dict()
        self.dictLocation=dict()
    
    def add_voiture(self,matricule,marque,couleur,etat,date,prix):
        v=Voiture(matricule,marque,couleur,etat,date,prix)
        self.dictVoiture[v.matricule]=v

    def delete_car_mat(self,matricule):
        l=[]
        del self.dictVoiture[matricule]
        for key,value in self.dictLocation.items():
            if(matricule==value.matricule):
                l.append(key)
        for i in l:
            del self.dictLocation[i]
    
    def delete_car_marque(self,marque):
        l=[]
        l1=[]
        for i,j in self.dictVoiture.items():
            if(j.marque)==marque:
                l.append(i)
        for key,value in self.dictLocation.items():
            if(marque==self.dictVoiture[value.matricule].marque):
                l1.append(key)
        for i in l1:
            del self.dictLocation[i]
        for i in l:
            del self.dictVoiture[i]
        return len(l)
    
    def delete_car_age(self):
        j=int(date.today().day)
        m=int(date.today().month)
        y=int(date.today().year)
        dateNow=QDate(y,m,j)
        l=[]
        for i,j in self.dictVoiture.items():
            if ((j.date).daysTo(dateNow))>3652.5:
                l.append(i)
        return l
        
    def update_car_price(self,matricule,prix):
        self.dictVoiture[matricule].prix=prix
    
    def update_car_color(self,matricule,couleur):
        self.dictVoiture[matricule].couleur=couleur

    def add_client(self,cin,nom,prenom,age,adresse,mail,tel):
        c=Client(cin,nom,prenom,age,adresse,mail,tel)
        self.dictClient[c.cin]=c
    
    def delete_client(self,cin):
        l=[]
        del self.dictClient[cin]
        for key,value in self.dictLocation.items():
            if(cin==value.cin):
                l.append(key)
        for i in l:
            del self.dictLocation[i]

    
    def update_client_Adresse(self,cin,adresse):
        self.dictClient[cin].adresse=adresse
    
    def update_client_tel(self,cin,tel):
        self.dictClient[cin].tel=tel

    def update_client_mail(self,cin,mail):
        self.dictClient[cin].mail=mail

    def add_location(self,numero,cin,matricule,date,duree,montant):
        l=Location(numero,cin,matricule,date,duree,montant)
        self.dictLocation[l.numero]=l

    def delete_location(self,numero):
        del self.dictLocation[numero]
    
    def update_location_date(self,numero,date):
        self.dictLocation[numero].date=date

    def update_location_duree(self,numero,duree):
        self.dictLocation[numero].duree=duree

    def save_car(self):
        with open('Voiture.csv','w', newline='') as output:
            writer = csv.writer(output,delimiter=';')
            for key, value in self.dictVoiture.items():
                fields = [str(key), str(value.marque), str(value.couleur), str(value.etat), (value.date).toString('dd/MM/yyyy'), str(value.prix)] 
                writer.writerow(fields)
            output.close()

    def save_client(self):
        with open('Client.csv','w', newline='') as output:
            writer = csv.writer(output,delimiter=';')
            for key, value in self.dictClient.items():
                fields = [str(key), str(value.nom), str(value.prenom), str(value.age), str(value.adresse), str(value.mail), str(value.tel)] 
                writer.writerow(fields)
            output.close()
    
    def save_Location(self):
        with open('Location.csv','w', newline='') as output:
            writer = csv.writer(output,delimiter=';')
            for key, value in self.dictLocation.items():
                fields = [str(key), str(value.cin), str(value.matricule), (value.date).toString("dd/MM/yyyy HH'h'"), str(value.duree), str(value.montant)] 
                writer.writerow(fields)
            output.close()
    
    def load_car(self):
        if (os.path.exists("Voiture.csv")):
            file = open("Voiture.csv")
            csvreader = csv.reader(file,delimiter=';')
            for row in csvreader:
                l=row[4].split('/')
                self.add_voiture(row[0],row[1],row[2],row[3],QDate(int(l[2]),int(l[1]),int(l[0])),row[5])
            file.close()
        else:
            with open('Voiture.csv','w', newline='') as output:
                output.close()

    def load_client(self):
        if (os.path.exists("Client.csv")):
            file = open("Client.csv")
            csvreader = csv.reader(file,delimiter=';')
            for row in csvreader:
                self.add_client(row[0],row[1],row[2],row[3],row[4],row[5],row[6])
            file.close()
        else:
            with open('Client.csv','w', newline='') as output:
                output.close()

    def load_Location(self):
        if (os.path.exists("Location.csv")):
            file = open("Location.csv")
            csvreader = csv.reader(file,delimiter=';')
            for row in csvreader:
                self.add_location(row[0],row[1],row[2],QDateTime.fromString(row[3],"dd/MM/yyyy HH'h'"),row[4],row[5])
            file.close()
        else:
            with open('Location.csv','w', newline='') as output:
                output.close()

    def list_client(self,view:QTableWidget):
        row=0
        view.setRowCount(len(self.dictClient))
        for key,item in self.dictClient.items():
            view.setItem(row,0,QTableWidgetItem(str(key)))
            view.setItem(row,1,QTableWidgetItem(str(item.nom)))
            view.setItem(row,2,QTableWidgetItem(str(item.prenom)))
            view.setItem(row,3,QTableWidgetItem(str(item.age)))
            view.setItem(row,4,QTableWidgetItem(str(item.adresse)))
            view.setItem(row,5,QTableWidgetItem(str(item.mail)))
            view.setItem(row,6,QTableWidgetItem(str(item.tel)))
            row=row+1
    
    def list_car(self,view:QTableWidget):
        row=0
        view.setRowCount(len(self.dictVoiture))
        for key,item in self.dictVoiture.items():
            view.setItem(row,0,QTableWidgetItem(str(key)))
            view.setItem(row,1,QTableWidgetItem(str(item.marque)))
            view.setItem(row,2,QTableWidgetItem(str(item.couleur)))
            view.setItem(row,3,QTableWidgetItem(str(item.etat)))
            view.setItem(row,4,QTableWidgetItem(str((item.date).toString('dd/MM/yyyy'))))
            view.setItem(row,5,QTableWidgetItem(str(item.prix)))
            row=row+1
    
    def search_client_cin(self,cin,nom,prenom,age,adresse,mail,tel):
        c=self.dictClient[cin]
        nom.setText(str(c.nom))
        prenom.setText(str(c.prenom))
        age.setText(str(c.age))
        adresse.setText(str(c.adresse))
        mail.setText(str(c.mail))
        tel.setText(str(c.tel))

    def search_car_matricule(self,matricule,marque,couleur,etat,date,prix):
        c=self.dictVoiture[matricule]
        marque.setText(str(c.marque))
        couleur.setText(str(c.couleur))
        etat.setText(str(c.etat))
        date.setDate(c.date)
        prix.setText(str(c.prix))

    def search_car_marque(self,marque,view:QTableWidget):
        row=0
        totalrow=0
        for key,item in self.dictVoiture.items():
            if(item.marque==marque):
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictVoiture.items():
            if(item.marque==marque):
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.marque)))
                view.setItem(row,2,QTableWidgetItem(str(item.couleur)))
                view.setItem(row,3,QTableWidgetItem(str(item.etat)))
                view.setItem(row,4,QTableWidgetItem(str((item.date).toString('dd/MM/yyyy'))))
                view.setItem(row,5,QTableWidgetItem(str(item.prix)))
                row=row+1
        
    def search_car_color(self,color,view:QTableWidget):
        
        row=0
        totalrow=0
        for key,item in self.dictVoiture.items():
            if(item.couleur==color):
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictVoiture.items():
            if(item.couleur==color):
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.marque)))
                view.setItem(row,2,QTableWidgetItem(str(item.couleur)))
                view.setItem(row,3,QTableWidgetItem(str(item.etat)))
                view.setItem(row,4,QTableWidgetItem(str((item.date).toString('dd/MM/yyyy'))))
                view.setItem(row,5,QTableWidgetItem(str(item.prix)))
                row=row+1

    def search_car_dispo(self,view:QTableWidget):
        row=0
        totalrow=0
        for key,item in self.dictVoiture.items():
            if(item.etat=="Disponible"):
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictVoiture.items():
            if(item.etat=="Disponible"):
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.marque)))
                view.setItem(row,2,QTableWidgetItem(str(item.couleur)))
                view.setItem(row,3,QTableWidgetItem(str(item.etat)))
                view.setItem(row,4,QTableWidgetItem(str((item.date).toString('dd/MM/yyyy'))))
                view.setItem(row,5,QTableWidgetItem(str(item.prix)))
                row=row+1
    
    def search_car_louee(self,view:QTableWidget):
        row=0
        totalrow=0
        for key,item in self.dictVoiture.items():
            if(item.etat=="Louée"):
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictVoiture.items():
            if(item.etat=="Louée"):
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.marque)))
                view.setItem(row,2,QTableWidgetItem(str(item.couleur)))
                view.setItem(row,3,QTableWidgetItem(str(item.etat)))
                view.setItem(row,4,QTableWidgetItem(str((item.date).toString('dd/MM/yyyy'))))
                view.setItem(row,5,QTableWidgetItem(str(item.prix)))
                row=row+1
    
    def calcul_montant(self,matricule,duree,montant):
        if(matricule.currentIndex()==-1 or matricule.currentText() not in self.dictVoiture.keys()):
            montant.setText("")
        else:
            m=int((self.dictVoiture[matricule.currentText()].prix))*int(duree.text())
            montant.setText(str(m)+ ' DH')
    def list_loc(self,view:QTableWidget):
        row=0
        view.setRowCount(len(self.dictLocation))
        for key,item in self.dictLocation.items():
            view.setItem(row,0,QTableWidgetItem(str(key)))
            view.setItem(row,1,QTableWidgetItem(str(item.cin)))
            view.setItem(row,2,QTableWidgetItem(str(item.matricule)))
            view.setItem(row,3,QTableWidgetItem(str((item.date).toString("dd/MM/yyyy HH'h'"))))
            view.setItem(row,4,QTableWidgetItem(str(item.duree)))
            view.setItem(row,5,QTableWidgetItem(str(item.montant)))
            row=row+1
    def search_loc_cin(self,cin,view:QTableWidget):
        row=0
        totalrow=0
        for key,item in self.dictLocation.items():
            if(item.cin==cin):
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictLocation.items():
            if(item.cin==cin):
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.cin)))
                view.setItem(row,2,QTableWidgetItem(str(item.matricule)))
                view.setItem(row,3,QTableWidgetItem(str((item.date).toString("dd/MM/yyyy HH'h'"))))
                view.setItem(row,4,QTableWidgetItem(str(item.duree)))
                view.setItem(row,5,QTableWidgetItem(str(item.montant)))
                row=row+1
    def search_loc_matricule(self,matricule,view:QTableWidget):
        row=0
        totalrow=0
        for key,item in self.dictLocation.items():
            if(item.matricule==matricule):
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictLocation.items():
            if(item.matricule==matricule):
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.cin)))
                view.setItem(row,2,QTableWidgetItem(str(item.matricule)))
                view.setItem(row,3,QTableWidgetItem(str((item.date).toString("dd/MM/yyyy HH'h'"))))
                view.setItem(row,4,QTableWidgetItem(str(item.duree)))
                view.setItem(row,5,QTableWidgetItem(str(item.montant)))
                row=row+1

    def imprimer_facture(self,numero):
        location=self.dictLocation[numero]
        voiture=self.dictVoiture[self.dictLocation[numero].matricule]
        client1=self.dictClient[self.dictLocation[numero].cin]
        dirname = os.path.dirname(__file__)
        inputfile = os.path.join(dirname, "Facture\\facture.xlsx")
        outputfile = os.path.join(dirname, "Facture\\"+str(numero)+".pdf")
        wb = load_workbook(inputfile)
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[0]]
        timenow=QDateTime.currentDateTime().toString("dd/MM/yyyy HH'h'")
        #Then update as you want it
        #Sheet1.cell(row = 14, column = 11).value = '14/5/2022' #This will change the cell(2,4) to 4
        Sheet1.cell(14,12).value=timenow
        Sheet1.cell(15,12).value=str(location.numero)
        Sheet1.cell(22,4).value=str(client1.nom)
        Sheet1.cell(23,4).value=str(client1.prenom)
        Sheet1.cell(24,4).value=str(client1.cin)
        Sheet1.cell(25,4).value=str(client1.age)+' ans'
        Sheet1.cell(26,4).value=str(client1.adresse)
        Sheet1.cell(27,4).value=str(client1.mail)
        Sheet1.cell(28,4).value=str(client1.tel)
        Sheet1.cell(32,10).value=str(location.duree)+' j'
        Sheet1.cell(32,11).value=str(voiture.prix)
        Sheet1.cell(32,12).value=str(location.montant)
        ch="Voiture "+str(voiture.marque)+" "+str(voiture.couleur)
        Sheet1.cell(32,3).value=ch
        Sheet1.cell(33,3).value=" Matricule N°: "+str(voiture.matricule)
        Sheet1.cell(34,3).value=" Date de Location: "+location.date.toString("dd/MM/yyyy HH'h'")
        wb.save(inputfile)
        wb.close()
        xlApp = client.Dispatch("Excel.Application")
        books = xlApp.Workbooks.Open(inputfile)
        ws = books.Worksheets[0]
        ws.Visible = 1
        ws.ExportAsFixedFormat(0, outputfile)
        books.Close()
        subprocess.Popen(outputfile, shell=True)
    def search_loc_date(self,date:QDate,view:QTableWidget):
        row=0
        totalrow=0
        for key,item in self.dictLocation.items():
            if((item.date).date()==date):
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictLocation.items():
            if((item.date).date()==date):
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.cin)))
                view.setItem(row,2,QTableWidgetItem(str(item.matricule)))
                view.setItem(row,3,QTableWidgetItem(str((item.date).toString("dd/MM/yyyy HH'h'"))))
                view.setItem(row,4,QTableWidgetItem(str(item.duree)))
                view.setItem(row,5,QTableWidgetItem(str(item.montant)))
                row=row+1
    def search_loc_duree(self,duree,view:QTableWidget):
        row=0
        totalrow=0
        for key,item in self.dictLocation.items():
            if(item.duree==duree):
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictLocation.items():
            if(item.duree==duree):
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.cin)))
                view.setItem(row,2,QTableWidgetItem(str(item.matricule)))
                view.setItem(row,3,QTableWidgetItem(str((item.date).toString("dd/MM/yyyy HH'h'"))))
                view.setItem(row,4,QTableWidgetItem(str(item.duree)))
                view.setItem(row,5,QTableWidgetItem(str(item.montant)))
                row=row+1

    def search_loc_2_date(self,date1,date2,view:QTableWidget):
        row=0
        totalrow=0
        for key,item in self.dictLocation.items():
            if((item.date).date()>=date1 and (item.date).date()<=date2):
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictLocation.items():
            if((item.date).date()>=date1 and (item.date).date()<=date2):
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.cin)))
                view.setItem(row,2,QTableWidgetItem(str(item.matricule)))
                view.setItem(row,3,QTableWidgetItem(str((item.date).toString("dd/MM/yyyy HH'h'"))))
                view.setItem(row,4,QTableWidgetItem(str(item.duree)))
                view.setItem(row,5,QTableWidgetItem(str(item.montant)))
                row=row+1

    def search_loc_2_date_car(self,date1,date2,view:QTableWidget):
        row=0
        totalrow=0
        l=[]
        for key,item in self.dictLocation.items():
            if((item.date).date()>=date1 and (item.date).date()<=date2):
                totalrow+=1
                l.append(item.matricule)
        view.setRowCount(totalrow)
        for i in l:
            item=self.dictVoiture[i]
            view.setItem(row,0,QTableWidgetItem(str(i)))
            view.setItem(row,1,QTableWidgetItem(str(item.marque)))
            view.setItem(row,2,QTableWidgetItem(str(item.couleur)))
            view.setItem(row,3,QTableWidgetItem(str(item.etat)))
            view.setItem(row,4,QTableWidgetItem(str((item.date).toString('dd/MM/yyyy'))))
            view.setItem(row,5,QTableWidgetItem(str(item.prix)))
            row=row+1

    def update_date_loc(self,numero,newdate):
        self.dictLocation[numero].date=newdate
    
    def update_duree_loc(self,numero,newduree):
        self.dictLocation[numero].duree=newduree
    
    def benefice_dash(self):
        dateNow=QDateTime.currentDateTime()
        Mois=dateNow.toString('MM/yyyy')
        benefice=0
        for value in self.dictLocation.values():
            if(value.date).toString('MM/yyyy')==Mois:
                montant=value.montant.replace('DH','')
                benefice=benefice+int(montant)
        return str(benefice)+' DH'
    
    def dispo_car(self):
        dispo=0
        for i in self.dictVoiture.values():
            if i.etat=="Disponible":
                dispo+=1
        return dispo
    
    def most_car_rent(self):
        dictloc=dict()
        for item in self.dictLocation.values():
            if(item.matricule in dictloc):
                dictloc[item.matricule]+=1
            else:
                dictloc[item.matricule]=1
        max=0
        if(len(dictloc)==0):
            return ''
        else:
            for key,value in dictloc.items():
                if(max<value):
                    max=value
                    matricule=key
            return self.dictVoiture[matricule].marque
    
    def most_client_rent(self):
        dictclient=dict()
        for item in self.dictLocation.values():
            if(item.cin in dictclient):
                dictclient[item.cin]+=1
            else:
                dictclient[item.cin]=1
        max=0
        if(len(dictclient)==0):
            return ''
        else:
            for key,value in dictclient.items():
                if(max<value):
                    max=value
                    cin=key
            return self.dictClient[cin].nom+' '+self.dictClient[cin].prenom

    def revenu_today(self):
        dateNow=QDateTime.currentDateTime()
        Mois=dateNow.toString('dd/MM/yyyy')
        benefice=0
        for value in self.dictLocation.values():
            if(value.date).toString('dd/MM/yyyy')==Mois:
                montant=value.montant.replace('DH','')
                benefice=benefice+int(montant)
        return str(benefice)+' DH'

    def list_cars_month(self,view:QTableWidget):
        dateNow=QDateTime.currentDateTime()
        Mois=dateNow.toString('MM/yyyy')
        row=0
        totalrow=0
        for value in self.dictLocation.values():
            if(value.date).toString('MM/yyyy')==Mois:
                totalrow+=1
        view.setRowCount(totalrow)
        for key,item in self.dictLocation.items():
            if(item.date).toString('MM/yyyy')==Mois:
                view.setItem(row,0,QTableWidgetItem(str(key)))
                view.setItem(row,1,QTableWidgetItem(str(item.cin)))
                view.setItem(row,2,QTableWidgetItem(str(item.matricule)))
                view.setItem(row,3,QTableWidgetItem(str((item.date).toString("dd/MM/yyyy HH'h'"))))
                view.setItem(row,4,QTableWidgetItem(str(item.duree)))
                row=row+1

    def refresh_rent(self,l:QLabel):
        dateToday=QDateTime.currentDateTime()
        dateNow=dateToday
        liste=[]
        for value in self.dictLocation.values():
            if((value.date)<=dateNow<=(value.date).addDays(int(value.duree))):
                liste.append(value.matricule)
                self.dictVoiture[value.matricule].etat="Louée"
            elif(not((value.date)<=dateNow<=(value.date).addDays(int(value.duree))) and value.matricule not in liste):
                self.dictVoiture[value.matricule].etat="Disponible"
        ch="       "+dateToday.toString('hh:mm:ss')
        l.setText(ch)

                